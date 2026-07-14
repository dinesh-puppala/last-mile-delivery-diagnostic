import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.chart import BarChart, PieChart, Reference

FONT = "Arial"
title_f   = Font(name=FONT, bold=True, size=14, color="1F3864")
sub_f     = Font(name=FONT, italic=True, size=10, color="595959")
hdr_f     = Font(name=FONT, bold=True, size=10, color="FFFFFF")
sec_f     = Font(name=FONT, bold=True, size=11, color="1F3864")
lbl_f     = Font(name=FONT, size=10, color="000000")
lblb_f    = Font(name=FONT, bold=True, size=10, color="000000")
input_f   = Font(name=FONT, size=9, color="0000FF")
formula_f = Font(name=FONT, size=10, color="000000")
mono_f    = Font(name=FONT, size=9, color="333333")
note_f    = Font(name=FONT, size=9, italic=True, color="808080")

hdr_fill   = PatternFill("solid", fgColor="1F3864")
sec_fill   = PatternFill("solid", fgColor="D9E1F2")
code_fill  = PatternFill("solid", fgColor="F2F2F2")
green_fill = PatternFill("solid", fgColor="E2EFDA")
red_fill   = PatternFill("solid", fgColor="FCE4E4")

thin = Side(style="thin", color="BFBFBF")
box  = Border(left=thin, right=thin, top=thin, bottom=thin)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
center = Alignment(horizontal="center", vertical="center")
right = Alignment(horizontal="right", vertical="center")

wb = load_workbook("/home/claude/proj/LastMile_Misdelivery_Model.xlsx")

# Load data
raw = pd.read_csv("/home/claude/proj/SAP_SD_Delivery_Export_RAW.csv")
clean = pd.read_csv("/home/claude/proj/Delivery_Data_CLEAN.csv")
region_carrier = pd.read_csv("/home/claude/proj/sql_failure_by_region_carrier.csv")
root_cause = pd.read_csv("/home/claude/proj/sql_root_cause.csv")
leadtime = pd.read_csv("/home/claude/proj/sql_leadtime_by_carrier.csv")

with open("/home/claude/proj/sql_headline_numbers.txt") as f:
    headline = dict(line.strip().split("=") for line in f if line.strip())

# =========================================================
# TAB: Raw ERP Export (sample)
# =========================================================
ws = wb.create_sheet("ERP Raw Export (Sample)")
ws.sheet_view.showGridLines = False
for c, w in {"A":2,"B":12,"C":12,"D":16,"E":18,"F":16,"G":20,"H":20,"I":20,"J":16,"K":14,"L":14}.items():
    ws.column_dimensions[c].width = w

ws["B2"] = "Simulated SAP SD Delivery Export — Raw Pull"; ws["B2"].font = title_f
ws["B3"] = "First 150 of 1,020 raw records as pulled from the ERP report, before cleaning. Note inconsistent date formats, casing, and duplicate rows — typical of a real export."
ws["B3"].font = sub_f
ws.merge_cells("B3:L3")

headers = list(raw.columns)
for i, h in enumerate(headers):
    c = ws.cell(row=5, column=2+i, value=h)
    c.font = hdr_f; c.fill = hdr_fill; c.alignment = center; c.border = box

sample = raw.head(150)
for r_idx, row in enumerate(sample.itertuples(index=False), start=6):
    for c_idx, val in enumerate(row):
        cell = ws.cell(row=r_idx, column=2+c_idx, value=val)
        cell.font = mono_f; cell.border = box; cell.alignment = Alignment(horizontal="left")

ws.freeze_panes = "B6"

# =========================================================
# TAB: Data Cleaning (Excel formulas)
# =========================================================
wc = wb.create_sheet("Data Cleaning (Excel)")
wc.sheet_view.showGridLines = False
for c, w in {"A":2,"B":16,"C":16,"D":18,"E":14,"F":16,"G":16,"H":18,"I":22}.items():
    wc.column_dimensions[c].width = w

wc["B2"] = "Data Cleaning — Excel Formulas"; wc["B2"].font = title_f
wc["B3"] = "TRIM/CLEAN standardize messy raw fields; IFERROR keeps output clean when data is missing. First 100 records shown."
wc["B3"].font = sub_f
wc.merge_cells("B3:I3")

clean_headers = ["Delivery_ID", "Region_RAW", "Region_CLEAN", "Carrier_RAW", "Carrier_CLEAN",
                  "Status_RAW", "Status_CLEAN", "Lead_Time_Days (NETWORKDAYS)"]
for i, h in enumerate(clean_headers):
    c = wc.cell(row=5, column=2+i, value=h)
    c.font = hdr_f; c.fill = hdr_fill; c.alignment = center; c.border = box

sample2 = raw.head(100).reset_index(drop=True)
for i in range(len(sample2)):
    r = 6 + i
    wc.cell(row=r, column=2, value=int(sample2.loc[i, "Delivery_ID"])).border = box
    wc.cell(row=r, column=2).font = lbl_f

    wc.cell(row=r, column=3, value=sample2.loc[i, "Region"]).border = box
    wc.cell(row=r, column=3).font = lbl_f

    # TRIM/CLEAN + PROPER to standardize casing/whitespace
    clean_cell = wc.cell(row=r, column=4, value=f'=IFERROR(PROPER(TRIM(CLEAN(C{r}))),"Unknown")')
    clean_cell.border = box; clean_cell.font = formula_f

    wc.cell(row=r, column=5, value=sample2.loc[i, "Carrier"]).border = box
    wc.cell(row=r, column=5).font = lbl_f

    carrier_clean = wc.cell(row=r, column=6, value=f'=IFERROR(PROPER(TRIM(CLEAN(E{r}))),"Unknown")')
    carrier_clean.border = box; carrier_clean.font = formula_f

    wc.cell(row=r, column=7, value=sample2.loc[i, "Delivery_Status"]).border = box
    wc.cell(row=r, column=7).font = lbl_f

    status_clean = wc.cell(row=r, column=8, value=f'=IFERROR(PROPER(TRIM(CLEAN(G{r}))),"Unknown")')
    status_clean.border = box; status_clean.font = formula_f

    # Lead time using NETWORKDAYS between ship and promised date parsed via DATEVALUE with IFERROR fallback
    lt_cell = wc.cell(row=r, column=9,
        value=f'=IFERROR(NETWORKDAYS(DATEVALUE("{sample2.loc[i,"Ship_Date"]}"),DATEVALUE("{sample2.loc[i,"Promised_Delivery_Date"]}")),"N/A")')
    lt_cell.border = box; lt_cell.font = formula_f; lt_cell.alignment = center

# =========================================================
# TAB: SQL Diagnostic Summary
# =========================================================
ws2 = wb.create_sheet("SQL Diagnostic Summary")
ws2.sheet_view.showGridLines = False
for c, w in {"A":2,"B":26,"C":18,"D":18,"E":18,"F":18,"G":40}.items():
    ws2.column_dimensions[c].width = w

ws2["B2"] = "SQL Diagnostic Analysis — Full Dataset (1,000 records)"; ws2["B2"].font = title_f
ws2["B3"] = "Queries run against the cleaned delivery dataset via SQL (SQLite). Results below are the actual query outputs."
ws2["B3"].font = sub_f
ws2.merge_cells("B3:G3")

# Headline KPIs
ws2["B5"] = "Headline KPIs"; ws2["B5"].font = sec_f
kpis = [
    ("Total records analyzed", headline["total_records"], "#,##0"),
    ("First-attempt delivery rate", float(headline["first_attempt_pct"])/100, "0.0%"),
    ("OTIF (On-Time, In-Full)", float(headline["otif_pct"])/100, "0.0%"),
]
r = 6
for lbl, val, fmt in kpis:
    ws2.cell(row=r, column=2, value=lbl).font = lbl_f
    c = ws2.cell(row=r, column=3, value=float(val)); c.number_format = fmt; c.font = lblb_f
    c.fill = green_fill; c.alignment = right; c.border = box
    ws2.cell(row=r, column=2).border = box
    r += 1

# SQL query text (documented)
r += 1
ws2.cell(row=r, column=2, value="Query used — First-Attempt Rate").font = sec_f
r += 1
sql1 = ("SELECT ROUND(100.0 * SUM(CASE WHEN Delivery_Status IN "
        "('Delivered','Delivered - Late') THEN 1 ELSE 0 END) / COUNT(*), 2) "
        "AS first_attempt_pct FROM deliveries;")
ws2.merge_cells(f"B{r}:G{r}")
c = ws2.cell(row=r, column=2, value=sql1); c.font = mono_f; c.fill = code_fill; c.alignment = left
ws2.row_dimensions[r].height = 30
r += 2

ws2.cell(row=r, column=2, value="Query used — OTIF Rate").font = sec_f
r += 1
sql2 = ("SELECT ROUND(100.0 * SUM(CASE WHEN Delivery_Status='Delivered' AND "
        "Order_Complete='Y' AND date(Actual_Delivery_Date)<=date(Promised_Delivery_Date) "
        "THEN 1 ELSE 0 END) / COUNT(*), 2) AS otif_pct FROM deliveries;")
ws2.merge_cells(f"B{r}:G{r}")
c = ws2.cell(row=r, column=2, value=sql2); c.font = mono_f; c.fill = code_fill; c.alignment = left
ws2.row_dimensions[r].height = 30
r += 2

# Failure by region x carrier table
ws2.cell(row=r, column=2, value="Failure Rate by Region \u00d7 Carrier (worst combinations)").font = sec_f
r += 1
fc_headers = ["Region", "Carrier", "Total Deliveries", "Failures", "Failure Rate %"]
for i, h in enumerate(fc_headers):
    c = ws2.cell(row=r, column=2+i, value=h); c.font = hdr_f; c.fill = hdr_fill; c.alignment = center; c.border = box
r += 1
fc_start = r
for _, row_data in region_carrier.iterrows():
    ws2.cell(row=r, column=2, value=row_data["Region"]).border = box
    ws2.cell(row=r, column=3, value=row_data["Carrier"]).border = box
    ws2.cell(row=r, column=4, value=int(row_data["total_deliveries"])).border = box
    ws2.cell(row=r, column=5, value=int(row_data["failures"])).border = box
    fr_cell = ws2.cell(row=r, column=6, value=float(row_data["failure_rate_pct"])/100)
    fr_cell.number_format = "0.0%"; fr_cell.border = box
    for col in range(2, 7):
        ws2.cell(row=r, column=col).font = lbl_f
        ws2.cell(row=r, column=col).alignment = center if col > 2 else Alignment(horizontal="left")
    r += 1
fc_end = r - 1

# conditional formatting: highlight high failure rate
ws2.conditional_formatting.add(f"F{fc_start}:F{fc_end}",
    ColorScaleRule(start_type="min", start_color="C6EFCE", end_type="max", end_color="FFC7CE"))

r += 1
ws2.cell(row=r, column=2, value="Root Cause Breakdown (of all failures)").font = sec_f
r += 1
rc_headers = ["Failure Reason", "Count", "% of Failures"]
for i, h in enumerate(rc_headers):
    c = ws2.cell(row=r, column=2+i, value=h); c.font = hdr_f; c.fill = hdr_fill; c.alignment = center; c.border = box
r += 1
rc_start_row = r
for _, row_data in root_cause.iterrows():
    ws2.cell(row=r, column=2, value=row_data["Failure_Reason_Code"]).border = box
    ws2.cell(row=r, column=3, value=int(row_data["failure_count"])).border = box
    pct_cell = ws2.cell(row=r, column=4, value=float(row_data["pct_of_failures"])/100)
    pct_cell.number_format = "0.0%"; pct_cell.border = box
    for col in range(2, 5):
        ws2.cell(row=r, column=col).font = lbl_f
        ws2.cell(row=r, column=col).alignment = center if col > 2 else Alignment(horizontal="left")
    r += 1
rc_end_row = r - 1

# Root cause pie chart
pie = PieChart()
pie.title = "Failure Root Cause Breakdown"
data = Reference(ws2, min_col=3, min_row=rc_start_row-1, max_row=rc_end_row)
cats = Reference(ws2, min_col=2, min_row=rc_start_row, max_row=rc_end_row)
pie.add_data(data, titles_from_data=True)
pie.set_categories(cats)
pie.height = 7; pie.width = 10
ws2.add_chart(pie, f"H6")

r += 1
ws2.cell(row=r, column=2, value="Lead Time Variance by Carrier (days, ship-to-delivery)").font = sec_f
r += 1
lt_headers = ["Carrier", "Avg Lead Time (days)", "Std Dev (days)", "Deliveries"]
for i, h in enumerate(lt_headers):
    c = ws2.cell(row=r, column=2+i, value=h); c.font = hdr_f; c.fill = hdr_fill; c.alignment = center; c.border = box
r += 1
for _, row_data in leadtime.iterrows():
    ws2.cell(row=r, column=2, value=row_data["Carrier"]).border = box
    ws2.cell(row=r, column=3, value=float(row_data["avg_lead_time_days"])).border = box
    ws2.cell(row=r, column=4, value=float(row_data["stdev_lead_time_days"])).border = box
    ws2.cell(row=r, column=5, value=int(row_data["n_deliveries"])).border = box
    for col in range(2, 6):
        ws2.cell(row=r, column=col).font = lbl_f
        ws2.cell(row=r, column=col).alignment = center if col > 2 else Alignment(horizontal="left")
    r += 1

# =========================================================
# UPDATE Dashboard: add operational diagnostic KPI cards
# =========================================================
wd = wb["Dashboard"]

wd["B22"] = "Operational Diagnostic — Simulated ERP Data (1,000 deliveries)"
wd["B22"].font = sec_f

def kpi(anchor, label, value, fmt, fill=green_fill):
    col = anchor[0]; row = int(anchor[1:])
    lc = wd[f"{col}{row}"]; lc.value = label
    lc.font = Font(name=FONT, bold=True, size=9, color="FFFFFF")
    lc.fill = hdr_fill; lc.alignment = center; lc.border = box
    vc = wd[f"{col}{row+1}"]; vc.value = value; vc.number_format = fmt
    vc.font = Font(name=FONT, bold=True, size=13, color="1F3864")
    vc.alignment = center; vc.fill = fill; vc.border = box

kpi("B24", "First-Attempt Rate", float(headline["first_attempt_pct"])/100, "0.0%")
kpi("D24", "OTIF Rate", float(headline["otif_pct"])/100, "0.0%")
kpi("F24", "Worst Failure Combo", float(headline["worst_combo_failure_rate"])/100, "0.0%", fill=red_fill)

wd["B27"] = f"Worst-performing combination: {headline['worst_combo_carrier']} in {headline['worst_combo_region']} region"
wd["B27"].font = note_f
wd.merge_cells("B27:G27")
wd["B28"] = f"Top failure cause: {headline['top_failure_cause']} ({headline['top_failure_cause_pct']}% of all failures)"
wd["B28"].font = note_f
wd.merge_cells("B28:G28")
wd["B29"] = "See 'SQL Diagnostic Summary' and 'Data Cleaning (Excel)' tabs for full methodology and query documentation."
wd["B29"].font = note_f
wd.merge_cells("B29:G29")

wb.save("/home/claude/proj/LastMile_Misdelivery_Model.xlsx")
print("saved")
